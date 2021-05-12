from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import sys
import glob
import shutil
import os
from datetime import datetime
from openpyxl import load_workbook

source_path = ".\Website_Automated_Upload/"
target_path = ".\Website_Automated_Processed/"
manual_path = ".\Website_Manual_Upload/"
today = datetime.today()
file_list = glob.glob(source_path + '\\*.xlsx')
row_count = []
primary_client_list = []
third_party_client_list = []
primary_count = 0
third_party_count = 0
id_error_exists = None

print("Starting Client Access Upload Automator Tool....")

chrome_path = "chromedriver.exe"
chrome_options = Options()
chrome_options.add_argument("--log-level=3")
driver = webdriver.Chrome(options=chrome_options, executable_path=chrome_path)

for file in file_list:
    try:
        print("")
        workbook = load_workbook(filename=str(file))
        print("Client Access Form Found - ", file)
        sheet = workbook.active
        # Exclude header
        row_count.append(2)
    except:
        print("***Upload File Not found. Exiting. ****")
        time.sleep(10)
        break
        sys.exit()

    while True:
        try:
            # Loop for Primary Clients
            if sheet.cell(row=row_count[-1], column=1).value == "" or sheet.cell(row=row_count[-1],
                                                                                 column=1).value is None:
                break
            else:
                e_last_name = sheet.cell(row=row_count[-1], column=1).value
                e_first_name = sheet.cell(row=row_count[-1], column=2).value
                e_email = sheet.cell(row=row_count[-1], column=3).value
                # e_email = e_email.strip()
                for i in range(5, 12):
                    if sheet.cell(row=row_count[-1], column=i).value != "":
                        primary_client_list.append(sheet.cell(row=row_count[-1], column=i).value)
                        primary_count = primary_count + 1
                    else:
                        break
                # Loop for Third-Party Clients
                for i in range(12, 19):
                    if sheet.cell(row=row_count[-1], column=i).value != "":
                        third_party_client_list.append(sheet.cell(row=row_count[-1], column=i).value)
                        third_party_count = third_party_count + 1
                    else:
                        break

                print("Creating User...")
                print(" Last Name: ", e_last_name)
                print(" First Name: ", e_first_name)
                print(" Email: ", e_email)
                print(" Accountholder Accounts: ", primary_client_list)
                print(" 3rd Party Accounts: ", third_party_client_list)

                for prim_id in primary_client_list:
                    if len(prim_id) != 6:
                        id_error_exists = True

                for third_id in third_party_client_list:
                    if len(third_id) != 6:
                        id_error_exists = True

                if e_last_name is None or e_first_name is None or e_email is None or id_error_exists == True:
                    print("A required field looks incorrect, please review- ", file,
                          ".  Moved to Manual Upload Folder.")
                    primary_count = 0
                    third_party_count = 0
                    primary_client_list.clear()
                    third_party_client_list.clear()
                    # shutil.copy(file, manual_path+"\CAD Client Access_"+e_last_name+".xlsx")
                    shutil.copy(file, manual_path)
                    os.remove(file)
                    print("---------------------------------------------------------------------")

                    break
                else:
                    print("Initial User Data looks good.")

                driver.get("https://appsqa.canaccord.com/canaccordloginqa/User/Create/CCC")
                time.sleep(2)
                email = driver.find_element_by_id("Email")
                email.click()
                email.send_keys(e_email)
                time.sleep(2)
                f_name = driver.find_element_by_id("FirstName")
                f_name.click()
                f_name.send_keys(e_first_name)
                time.sleep(2)
                l_name = driver.find_element_by_id("LastName")
                l_name.click()
                l_name.send_keys(e_last_name)
                time.sleep(2)
                save_button = driver.find_element_by_xpath(
                    "/html/body/div/div[2]/div/div/div/div[2]/div/div/div/div[2]/form/div[4]/button")
                time.sleep(2)
                save_button.click()
                time.sleep(2)
                print("Inserting Accountholder Client IDs..")
                # Re_Direction to Add Client ID Page.
                for i in range(0, primary_count):
                    add_client_id = driver.find_element_by_xpath(
                        "/html/body/div/div[2]/div/div/div/div[2]/div/div/div[3]/div[1]/div[2]/div/a")
                    add_client_id.click()

                    # Re Direct to Add Client ID Page (2).
                    time.sleep(2)

                    client_id = driver.find_element_by_id("ClientId")
                    client_id.click()
                    client_id.send_keys(primary_client_list[i])

                    client_type = driver.find_element_by_id("Role")
                    client_type.send_keys("Primary")

                    client_save_button = driver.find_element_by_xpath(
                        "/html/body/div/div[2]/div/div/div/div[2]/div/div/div/div[2]/form/div[3]/button")
                    client_save_button.click()
                    time.sleep(2)

                print("Inserting any Third Party Client IDs..")
                for i in range(0, third_party_count):
                    add_client_id = driver.find_element_by_xpath(
                        "/html/body/div/div[2]/div/div/div/div[2]/div/div/div[3]/div[1]/div[2]/div/a")
                    add_client_id.click()

                    # Re Direct to Add Client ID Page (2).
                    time.sleep(2)

                    client_id = driver.find_element_by_id("ClientId")
                    client_id.click()
                    client_id.send_keys(third_party_client_list[i])

                    client_type = driver.find_element_by_id("Role")
                    client_type.send_keys("ThirdParty")

                    client_save_button = driver.find_element_by_xpath(
                        "/html/body/div/div[2]/div/div/div/div[2]/div/div/div/div[2]/form/div[3]/button")
                    client_save_button.click()
                    time.sleep(2)

                row_count.append(row_count[-1] + 1)
                primary_count = 0
                third_party_count = 0
                primary_client_list.clear()
                third_party_client_list.clear()

                file_time = str(today.strftime("%H.%M.%S"))
                shutil.copy(file, target_path + "\CAD Client Access_" + e_last_name + e_first_name + "_" + str(
                    today.strftime("%m%d%Y")) + "_" + file_time + ".xlsx")
                os.remove(file)
                print("Upload completed. Client Access Excel Form moved to Processed folder.")
                print("---------------------------------------------------------------------")
                time.sleep(1)
        except Exception as e:
            print(str(e))
            print("Error Found in File -", file, ".  Client Access form moved to Manual Upload Folder.")
            primary_count = 0
            third_party_count = 0
            primary_client_list.clear()
            third_party_client_list.clear()
            # shutil.copy(file, manual_path+"\CAD Client Access_"+e_last_name+e_first_name+".xlsx")
            shutil.copy(file, manual_path)
            os.remove(file)
            print("---------------------------------------------------------------------")
            break

print("")
print("All Client Forms have been successfully uploaded or moved into Manual Upload folder.")
print("")
driver.get("https://appsqa.canaccord.com/CanaccordLoginQA/User")
time.sleep(3)
