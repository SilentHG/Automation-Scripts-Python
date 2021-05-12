from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
import os
import subprocess
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
import datetime
import math

workbook = load_workbook(filename="Sample Upload Final.xlsx")
sheet = workbook.active
# assuming header is provided
print("No of rows in spreadsheet excluding header:", sheet.max_row - 1)
iterations = sheet.max_row

row_count = []
row_count.append(2)

# path = r"C:\x"
# os.system('cmd /c "chrome.exe --remote-debugging-port=9222 --user-data-dir="'+path)
# time.sleep(2)


chrome_path = r"C:\Users\shaki\PycharmProjects\_ScrappingModule\chromedriver.exe"

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = webdriver.Chrome(options=chrome_options, executable_path=chrome_path)

#driver.get('https://portal.dtcc.com')
#time.sleep(2)

# try to get the username input element until the page loads
#userInput = driver.find_element_by_xpath("//p[@id='userPass']/input[@name='username']")

# fill in username and password
#userInput.send_keys('')
#time.sleep(2)
#passInput = driver.find_element_by_xpath("//p[@id='userPass']/input[@name='password']")
#passInput.send_keys('')

# click on login button
#loginBtn = driver.find_element_by_xpath("//p[@id='Logimg']/input[@name='Login']")
#loginBtn.click()
# 
# 
#smarttrack buy-ins home page
#driver.get('https://portal.dtcc.com/cre5/bin/toHomeNonCNSAction.do')
#time.sleep(2)
#driver.get('https://portal.dtcc.com/cre5/bin/toCreateNonCNSRetransmittalAction.do?statusCd=NR')


# =============================================================================

while iterations > 1:
    try:

        security_id = sheet.cell(row=row_count[-1], column=3).value
        rece_part_id = sheet.cell(row=row_count[-1], column=2).value
        rece_part_id = rece_part_id[-4:]
        quantity = abs(float(sheet.cell(row=row_count[-1], column=5).value))
        frac_quantity, whole_quantity = math.modf(quantity)

        settlement_amount = abs(float(sheet.cell(row=row_count[-1], column=6).value))
        print(settlement_amount)
        frac_settle_amount, whole_settlement = math.modf(settlement_amount)
        print(frac_settle_amount,whole_settlement)
        settlement_date = str(sheet.cell(row=row_count[-1], column=7).value)
        print(settlement_date)
        settlement_date = datetime.datetime.strptime(settlement_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y')

        delivery_date = str(sheet.cell(row=row_count[-1], column=8).value)
        print(delivery_date)
        delivery_date = datetime.datetime.strptime(delivery_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y')

        org_id = driver.find_element_by_name("orgPartId1")
        org_id.click()
        org_id.send_keys("5099")
        time.sleep(0.5)

        sec_id = driver.find_element_by_xpath(
            "/html/body/table/tbody/tr[3]/td/table/tbody/tr[5]/td/form/table/tbody/tr[4]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/input")
        sec_id.click()
        sec_id.send_keys(security_id)
        time.sleep(0.5)
        rec_participant = driver.find_element_by_name("partRcvrId")
        rec_participant.click()
        rec_participant.send_keys(rece_part_id)
        time.sleep(0.5)
        exg_mark = driver.find_element_by_name("mktCd")
        exg_mark.send_keys("OTC")
        time.sleep(0.5)
        Quan_Whole = driver.find_element_by_name("shareQtWhole")
        Quan_Whole.click()
        Quan_Whole.send_keys(int(whole_quantity))
        time.sleep(0.5)
        Quan_Frac = driver.find_element_by_name("shareQtFrac")
        Quan_Frac.click()
        Quan_Frac.send_keys(str(frac_quantity)[2:])
        time.sleep(0.5)
        settlement_whole = driver.find_element_by_name("totalAmWhole")
        settlement_whole.click()
        settlement_whole.send_keys(int(whole_settlement))
        time.sleep(0.5)
        settlement_frac = driver.find_element_by_name("totalAmFrac")
        settlement_frac.click()
        settlement_frac.send_keys(str(frac_settle_amount)[2:])
        time.sleep(0.5)
        settlement_date_box = driver.find_element_by_name("settleDt")
        settlement_date_box.send_keys(settlement_date)
        time.sleep(0.5)
        delivery_date_box = driver.find_element_by_name("dlvryDt")
        delivery_date_box.send_keys(delivery_date)
        time.sleep(0.5)
        balance_order = driver.find_element_by_name("balOderCntlNo")

        #print("FLAG2")
        row_count.append(row_count[-1] + 1)
        iterations = iterations - 1
        time.sleep(3)

        sec_id.clear()
        time.sleep(0.5)
        rec_participant.clear()
        time.sleep(0.5)
        Quan_Whole.clear()
        time.sleep(0.5)
        Quan_Frac.clear()
        time.sleep(0.5)
        settlement_whole.clear()
        time.sleep(0.5)
        settlement_frac.clear()
        time.sleep(0.5)
        settlement_date_box.clear()
        time.sleep(0.5)
        delivery_date_box.clear()
    except Exception as e:
        print(" ERROR")
        print(str(e))
        break
